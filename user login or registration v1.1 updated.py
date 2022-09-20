
"""
Created on Tue Sep 18 12:06:40 2022

@author: mogith
"""
import pandas as pd
import openpyxl
from openpyxl import load_workbook

def passwordcheck(pwd): #function to check whether password satisfies all conditions
 accept_pwd=0     #password pointer 
 if (len(pwd)>=5 or len(pwd)<=16): #condition 1 checking length of passwd
    accept_pwd+=1
 else:
    print('password length must be min 5 and max 16')
 if (any(map(str.isdigit,pwd))==True): # checking condition whether passed consists of one integer
    accept_pwd+=1
 else:
    print("Password must have atleast one integer")
 if accept_pwd>1: #condition 3 password consists of atleast one upper case and lowercase letter
    uppercount=0 #counter for caps letter 
    lowercount=0 #counter for small letter
    for i in pwd:
        if i.isupper():  #checks whether password consists of atleast one capital letter
            uppercount+=1
            break
    else:
          print("Password must have one capital letter")
    for i in pwd:
        if i.islower():  #checks whether password consists of aleast one small letter
            lowercount+=1
            break
    else:
          print("Password must have one small letter")      
    if uppercount>=1 and lowercount>=1: 
      accept_pwd+=1
 if accept_pwd>2: # condition 4 executes only when above conditions met
    for i in range(len(pwd)):
        if pwd[i] in '[@_!#$%^&*()<>?/\|}{~:]':
            accept_pwd+=1

 if accept_pwd>3: #return value only when all 4 conditions met 
    return(pwd) 
 else:
     print("Enter valid password")
def registrationemail(email_ID): #function for registering email and check if satisfies all conditions 
    accept_email=True
    emailheader=(email_ID[0:(email_ID.index('@'))]) 
    if email_ID[0]=='@' : #email should not start with @
        accept_email=False
    elif (email_ID.index('.')-email_ID.index('@'))<2: # After @ email should have a domain name before '.'
        accept_email=False
    elif (accept_email==True):
        for i in range(len(emailheader)): ##header should not have any special integers or numbers
            if emailheader[i] in '1234567890':
               accept_email=False
               break
            elif emailheader[i] in '[@_!#$%^&*()<>?/\|}{~:]' :
                accept_email=False
                break
    if accept_email!=True: #if all conditions are not satisfied
        print("Enter valid Email address")
        return (0)
    else: # If all conditions are satisfied it returns email-ID
        return(email_ID)
     
def registerprocess(email,passwd): #main function which returns the mail id and password in tuple 
    if registrationemail(email) and passwordcheck(passwd): #only if both returns a value
        return(registrationemail(email),passwordcheck(passwd))
    else:
        print("enter valid registration email /password")

log_or_reg=input("Login or Registration(press 1 for reg & 2 for log):") # Asking user whether login or registration
if log_or_reg=='1':  #validating user choice 
    email=input('enter email address to register:') #if user chooses reg it will trigger registerprocess function
    details=(registerprocess(email)) #store the value as a tuple
    wb = openpyxl.load_workbook("./Userdatabase.xlsx") #uses openpyxl to load the workbook
    sh = wb['data']   #assigning sheet value sh to sheet name
    sh.append(details) #appending details which is in tuples format into new row
    wb.save("./Userdatabase.xlsx") #saving file
elif log_or_reg=='2':
     Emaillogin=input('enter email for logging in:') #if user chooses for login
     passwdlogin=input('enter password for logging in:')
     Df=pd.read_excel('./Userdatabase.xlsx') #reading excel file as dataframe
     Df_to_dict = Df.set_index('email_ID').T.to_dict() #converting Df into dict in order to query details as key and value
     if not (Df['email_ID'].str.contains(Emaillogin).any()): #if enetered email is not there in any of the columns/row
         print("login unsuccessful kindly check login credentials")
     elif(( Df_to_dict[Emaillogin]['password'])==passwdlogin ): #if user entered credentials where matching - bingo :) login success
         print('login successful')
     else:
         print(' password do not match the given email') #if password doesn't match the given email
         nextattempt=input('For Registration: press 1/forget_password: press 2/change_passsword:press 3=')  
         if nextattempt=='1': #will allow user to register again
             newreg=input("enter new email_id to register:") 
             newregemail=input("enter new password to register")
             newdetails=registerprocess(newreg,newregemail)
             wb = openpyxl.load_workbook("./Userdatabase.xlsx")
             sh = wb['data']
             sh.append(newdetails)
             wb.save("./Userdatabase.xlsx")
             print("you have sucessfully registered")
         elif nextattempt=='2': #will fetch the corresponding password for the email
             print("Password for your account is:" ,Df_to_dict[Emaillogin]['password'])
         elif nextattempt=='3': #will allow the user to change the password for the given email a
             newpassword=input("Enter new password for the account: ")
             newpasswordcheck=passwordcheck(newpassword)
             
             if (newpasswordcheck)!=None: #updates the given password in the excel file only if the passwordcheck function passes and validated
                 
                 Df.replace(Df_to_dict[Emaillogin]['password'],newpasswordcheck,inplace=True)
                 ExcelWorkbook = load_workbook('./Userdatabase.xlsx')
                 writer = pd.ExcelWriter('./Userdatabase.xlsx', engine = 'openpyxl',if_sheet_exists='replace',mode='a')
                 writer.book = ExcelWorkbook
                 writer.sheets = dict((ws.title, ws) for ws in ExcelWorkbook.worksheets)
                 Df.to_excel(writer, sheet_name='data', index=False)
                 writer.save()
                 writer.close()
                 print("New passsword successfully changed")
